from io import BytesIO
from typing import Hashable
from zipfile import BadZipFile

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from validation.decorators import validator
from validation.exceptions import ValidationException


def validate_worksheet_has_content(worksheet: Worksheet):
    """
    Validates that worksheet has content inside. It is checked by
    rows quantity.
    """
    worksheet_rows_quantity = worksheet.max_row - worksheet.min_row
    if worksheet_rows_quantity <= 1:
        raise ValidationException(
            "Worksheet rows quantity must be greater than 1. Not %s." %
            worksheet_rows_quantity
        )


def validate_worksheet_starts_from_left_top_corner(worksheet: Worksheet):
    """
    Validates that worksheet starts from the A1 cell.
    """
    if worksheet.min_column != 1 or worksheet.min_row != 1:
        raise ValidationException(
            "The worksheet must start from the A1 cell."
        )


def validate_worksheet_columns_size(worksheet: Worksheet, expected_columns_quantity: int):
    """
    Validates that the worksheet has fixed quantity of columns. If
    not, raises an exception.
    """
    max_columns_quantity = worksheet.max_column
    if max_columns_quantity != expected_columns_quantity:
        raise ValidationException(
            f"Worksheet must have exactly {expected_columns_quantity} columns."
            f" But has {max_columns_quantity}"
        )


def validate_rows_dont_have_duplicates():
    """
    Validates if all rows are unique. Otherwise, raises an exception.
    """

    existing_rows = set()

    @validator(name='validate_rows_dont_have_duplicates')
    def inner(record_row: Hashable, row_index: int):
        if record_row in existing_rows:
            raise ValidationException(
                f"Row on {row_index} is a duplicate of some previous record."
            )

    return inner


def validate_content_is_workbook(content: bytes) -> Workbook:
    """
    Validates that downloaded content from the file link is an Excel file.
    """
    try:
        return load_workbook(BytesIO(content))
    except BadZipFile as e:
        raise ValidationException("Downloaded content isn't a workbook.", log_msg=str(e))
