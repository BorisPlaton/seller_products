import pytest
from openpyxl.workbook import Workbook
from pydantic import BaseModel


@pytest.fixture
def workbook():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'record_id'
    ws['B1'] = 'record_name'
    ws['C1'] = 'record_active'
    ws['A2'] = 1
    ws['A3'] = 3
    ws['B2'] = 'string'
    ws['B3'] = 'another string'
    ws['C2'] = True
    ws['C3'] = False
    return wb


@pytest.fixture
def worksheet_record_class():
    class Record(BaseModel):
        record_id: int
        record_name: str
        record_active: bool

        def __hash__(self):
            return hash((type(self),) + tuple(i for i in self.__dict__))

    return Record
